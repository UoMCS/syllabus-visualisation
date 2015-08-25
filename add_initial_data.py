#!/usr/bin/python

from __future__ import print_function

import sys
import os
import glob
import itertools
import argparse

from flask import Flask
import openpyxl
import requests

from server.models import *
from server.api import get_categories

# Expected spreadsheet headers
HEADERS = [('A', 'Concept'),
            ('B', 'Name'), 
            ('C', 'Unit Code'),
            ('D', 'Taught'),
            ('E', 'Applied'),
            ('F', 'Assessed'),
            ('G', 'Context')]

def to_bool(val):
    return (1 if (val == 'y') else 0)

def is_custom(title):
    return title.startswith('UOM:')

def is_section(title):
    return "#" in title 

def normalize_custom_title(title):
    """ Applies WP-like normalization to a custom topics title """
    return title.replace('_', ' ')

def normalize_wp_title(title):
    """ Applies WP normalization to a title, so we get it's canonical form"""
    params = {
        'action': 'query',
        'titles': title,
        'format': 'json',
        'indexpageids': True
    }

    r = requests.get('http://en.wikipedia.org/w/api.php', params=params)
    responce = r.json()

    pageid = responce['query']['pageids'][0]

    assert pageid != '-1', 'Title not found'

    return responce['query']['pages'][pageid]['title']

def normalize_title(title):
    """ Applies correct type of normalization depending on topic type """
    if is_custom(title):
        return normalize_custom_title(title[4:])
    else:
        return normalize_wp_title(title)

def process_workbook(workbook):
    sheet = openpyxl.load_workbook(workbook).get_active_sheet()

    for header in HEADERS:
        if sheet.cell('%s1' % header[0]).value != header[1]:
            print("Error : Invalid cell in spreadsheet header cell %s" % header[0])
            sys.exit(1)

    # We couldn't add contexts straight away, as corresponding topics might not
    # yet be added. So we parse and save them here, and add after the topics are
    # added
    topic_contexts = [] 

    for row in range(2,sheet.get_highest_row()+1):
        if (sheet.cell('A%d' % row).value):
            concept = dict()

            for field in HEADERS:
                concept[field[1]] = sheet.cell('%s%d' % (field[0], row)).value
            
            # FIXME: Skipping sections for now
            if is_section(concept['Concept']):
                continue
            
            # Before topic title is normalized - determine if it's custom
            is_custom_concept = is_custom(concept['Concept'])

            # Name might be just a duplicated identifier - we don't need it then
            if concept['Name'] == concept['Concept']:
                concept['Name'] = None 

            concept['Concept'] = normalize_title(concept['Concept'])

            # Name might also be a normalized identifier - we don't need it either
            if concept['Name'] == concept['Concept']:
                concept['Name'] = None

            topic = None

            if is_custom_concept:
                topic = db.session.query(CustomTopic).filter_by(name=concept['Concept']).first()

                if not topic: # New topic
                    topic = CustomTopic(concept['Concept'])
                    topic.description = 'Added from spreadsheets'
                    db.session.add(topic)
            else:
                topic = db.session.query(Topic).filter_by(name=concept['Concept']).first()

                if not topic: # New topic
                    topic = Topic(concept['Concept'])
                    topic.categories = get_categories(topic)
                    db.session.add(topic)

            unit = db.session.query(Unit).filter_by(code=concept['Unit Code']).one()

            unit_topic = UnitTopic(unit.id, topic.id)
            unit_topic.alias = concept['Name']
            unit_topic.is_taught = to_bool(concept['Taught'])
            unit_topic.is_assessed = to_bool(concept['Assessed'])
            unit_topic.is_applied = to_bool(concept['Applied'])

            db.session.add(unit_topic)
            db.session.flush() # So that unit_topic.id is populated

            print(u'{} {}'.format(topic.id,topic.name))

            if concept['Context']:
                contexts = concept['Context'].split()

                # FIXME: Remove sections for now
                contexts = itertools.ifilterfalse(is_section, contexts)

                # Normalise titles
                contexts = map(normalize_title, contexts)

                topic_contexts.append((unit_topic.id, contexts))

            # Some lazy progress reporting
            sys.stdout.write('.')
            sys.stdout.flush()

    for unit_topic_id, contexts in topic_contexts:
        unit_topic = db.session.query(UnitTopic).filter_by(id=unit_topic_id).one()
        unit_topic.contexts = db.session.query(Topic).filter(Topic.name.in_(contexts)).all()

    print('Done')

def insert_units(units_filename):
    with open(units_filename) as f:
        for unit_line in f:
            code, name = map(str.strip, unit_line.split(',', 1))
            unit = Unit(code,name) 
            db.session.add(unit)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Insert initial data from spreadsheets')

    parser.add_argument('db_uri', help='Database URI to insert to (e.g. sqlite:///test.db)')
    parser.add_argument('data_dir', help='Directory with initial data')
    args = parser.parse_args()

    app = Flask(__name__)
    app.config['SQLALCHEMY_DATABASE_URI'] = args.db_uri 

    db.init_app(app)

    with app.app_context():
        db.create_all()

        units_filename = os.path.join(args.data_dir, 'units.txt')
        insert_units(units_filename)
        db.session.commit()

        spreadsheets = glob.glob(os.path.join(args.data_dir, '*.xlsx')) 
        for workbook in spreadsheets:
            print('Processing ' + workbook)
            process_workbook(workbook)
            db.session.commit()

